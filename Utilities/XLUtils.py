import openpyxl


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    return (sheet.max_row)

def columnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    return (sheet.max_cloumn)


def readdata(file, sheetName, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    return sheet.cell(rownum, columnnum).value


def writedata(file, sheetName, rownum, columnnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    sheet.cell(rownum, columnnum).value = data
    workbook.save(file)


# def fillGreencolor(file, sheetName, rownum, columnnum):
#     workbook = load_workbook(file)
#     sheet = workbook.active
#     greenfill = PatternFill(start_color='60b212', end_color='60b212', fill_type='solid')
#     sheet.cell(rownum, columnnum).fill = greenfill
#     workbook.save(file)
#
#
# def fillredncolor(file, sheetName, rownum, columnnum):
#     workbook = load_workbook(file)
#     sheet = workbook.active
#     redfill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
#     sheet.cell(rownum, columnnum).fill = redfill
#     workbook.save(file)

